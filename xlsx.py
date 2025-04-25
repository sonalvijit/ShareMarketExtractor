import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.worksheet.page import PageMargins
import time
from models import get_share_data

def save(data: dict):
    # Flatten the nested data structure into a list of rows
    rows = []
    for stock, details in data[0].items():
        for price_entry in details["prices"]:
            rows.append({
                "stock": stock,
                "stock_price": price_entry["share_value"],
                "p_change": price_entry["share_pts"],
                "G_N_L": price_entry["G_N_L"]
            })

    # Convert the rows into a DataFrame
    df = pd.DataFrame(rows)

    # Split the data into two halves
    mid_index = len(df) // 2
    first_half = df.iloc[:mid_index]
    second_half = df.iloc[mid_index:]

    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    # Set page margins (narrower for fitting)
    ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.2, bottom=0.2)

    # Add headers for both sections
    headers = df.columns.tolist()
    ws.append(headers + [""] + headers)  # Add a gap between the two sections

    # Define styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    consolas_font = Font(name="Consolas")

    # Add data rows for both sections
    for i in range(max(len(first_half), len(second_half))):
        row1 = first_half.iloc[i].tolist() if i < len(first_half) else [""] * len(headers)
        row2 = second_half.iloc[i].tolist() if i < len(second_half) else [""] * len(headers)
        ws.append(row1 + [""] + row2)  # Add a gap between the two sections

        last_row = ws.max_row

        # Apply green fill if Gain/Loss is "GAIN" in the first section
        if row1[3] == "GAIN":
            for col_idx in range(3, len(headers) + 1):
                ws.cell(row=last_row, column=col_idx).fill = green_fill

        # Apply green fill if Gain/Loss is "GAIN" in the second section
        if row2[3] == "GAIN":
            for col_idx in range(len(headers) + 4, len(headers) * 2 + 2):
                ws.cell(row=last_row, column=col_idx).fill = green_fill

    # Apply font and border to all filled cells
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                cell.font = consolas_font
                cell.border = thin_border

    # Add footer with the date
    ws.oddFooter.center.text = data[1]

    # Save the final Excel file
    excel_path_final = f"./stock_data_{data[1].replace("-","_")}.xlsx"
    wb.save(excel_path_final)

    print("Excel file generated:", excel_path_final)

# a = get_share_data("04-04-2025")
# save(a)